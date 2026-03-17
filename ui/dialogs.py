"""
ui/dialogs.py
All @st.dialog popups:
  - show_eye_popup         — cell-view with Excel highlight
  - show_field_history_dialog — per-field edit timeline
  - show_settings_dialog   — conf threshold + schema manager
  - show_schema_fields_dialog — required / accepted / custom field viewer
"""

import csv
import datetime

import streamlit as st
from openpyxl.utils import get_column_letter
from PIL import ImageDraw

from modules.field_history import _get_field_history
from modules.excel_renderer import (
    render_excel_sheet, get_cell_pixel_bbox, crop_context,
)


# ── Eye popup ─────────────────────────────────────────────────────────────────

@st.dialog("Cell View", width="large")
def show_eye_popup(field: str, info: dict, excel_path: str, sheet_name: str) -> None:
    import os
    st.markdown(f"### 📍 {field}")
    value      = info.get("modified", info["value"])
    target_row = info.get("excel_row")
    target_col = info.get("excel_col")

    col_a, col_b = st.columns([1, 1])
    with col_a:
        st.markdown("**Extracted Value**")
        st.code(value if value else "(empty)")
    with col_b:
        col_letter = get_column_letter(target_col) if target_col else "?"
        st.markdown(
            f"<div style='padding:10px 0;color:var(--t2);font-size:var(--sz-body);font-family:var(--font);'>"
            f"Cell: <span style='color:var(--blue);font-weight:bold;'>{col_letter}{target_row or '?'}</span>"
            f" &nbsp;|&nbsp; Row <span style='color:var(--t0);'>{target_row or '?'}</span>"
            f" · Col <span style='color:var(--t0);'>{target_col or '?'}</span></div>",
            unsafe_allow_html=True,
        )

    if not target_row or not target_col:
        st.warning("No cell location recorded for this field.")
        return

    ext = os.path.splitext(excel_path)[1].lower()

    if ext == ".csv":
        st.markdown("---")
        try:
            with open(excel_path, "r", encoding="utf-8-sig") as f:
                all_rows = list(csv.reader(f))
            if not all_rows:
                return
            n_rows  = len(all_rows)
            n_cols  = max(len(r) for r in all_rows)
            r0, r1  = max(0, target_row - 6), min(n_rows, target_row + 5)
            col_headers = "".join(
                f"<th style='background:var(--s0);color:var(--t3);font-size:var(--sz-xs);"
                f"padding:4px 8px;border:1px solid var(--b0);font-family:var(--mono);font-weight:600;'>"
                f"{get_column_letter(c + 1)}</th>"
                for c in range(n_cols)
            )
            thead = (
                f"<thead><tr>"
                f"<th style='background:var(--s0);color:var(--t3);font-size:var(--sz-xs);"
                f"padding:4px 8px;border:1px solid var(--b0);font-family:var(--mono);font-weight:600;'>#</th>"
                f"{col_headers}</tr></thead>"
            )
            tbody = ""
            for r_idx in range(r0, r1):
                row_data = all_rows[r_idx] if r_idx < len(all_rows) else []
                is_tr    = (r_idx + 1 == target_row)
                rn_style = (
                    "background:#1a2540;color:#4f9cf9;font-weight:bold;"
                    if is_tr
                    else "background:var(--s0);color:var(--t3);"
                )
                cells = (
                    f"<td style='{rn_style}font-size:var(--sz-xs);padding:4px 8px;"
                    f"border:1px solid var(--b0);text-align:center;font-family:var(--mono);'>"
                    f"{r_idx + 1}</td>"
                )
                for c_idx in range(n_cols):
                    cell_val = row_data[c_idx] if c_idx < len(row_data) else ""
                    is_tc    = is_tr and (c_idx + 1 == target_col)
                    if is_tc:
                        style = "background:#2a2010;border:2px solid var(--yellow);color:var(--t0);font-weight:bold;"
                    elif is_tr:
                        style = "background:var(--blue-g);border:1px solid rgba(79,156,249,0.2);color:var(--t1);"
                    else:
                        style = "background:var(--surface);border:1px solid var(--b0);color:var(--t2);"
                    cells += (
                        f"<td style='{style}font-size:var(--sz-sm);padding:5px 10px;"
                        f"max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;"
                        f"font-family:var(--font);'>{cell_val}</td>"
                    )
                tbody += f"<tr>{cells}</tr>"
            st.markdown(
                f"<div style='overflow-x:auto;border-radius:var(--radius);border:1px solid var(--b0);'>"
                f"<table style='border-collapse:collapse;width:100%;font-family:var(--font);'>"
                f"{thead}<tbody>{tbody}</tbody></table></div>",
                unsafe_allow_html=True,
            )
        except Exception as e:
            st.error(f"CSV preview error: {e}")
        return

    st.markdown("---")
    cache_key = f"_rendered_{excel_path}_{sheet_name}"
    with st.spinner("Rendering sheet…"):
        if cache_key not in st.session_state:
            rendered_img, col_starts, row_starts, merged_master = render_excel_sheet(
                excel_path, sheet_name, scale=1.0
            )
            st.session_state[cache_key] = (rendered_img, col_starts, row_starts, merged_master)
        else:
            rendered_img, col_starts, row_starts, merged_master = st.session_state[cache_key]

    try:
        img  = rendered_img.copy()
        draw = ImageDraw.Draw(img, "RGBA")
        x1, y1, x2, y2 = get_cell_pixel_bbox(col_starts, row_starts, target_row, target_col, merged_master)
        draw.rectangle([x1 + 1, y1 + 1, x2 - 1, y2 - 1], fill=(255, 230, 0, 80))
        draw.rectangle([x1, y1, x2, y2], outline=(245, 158, 11, 255), width=3)
        draw.rectangle([x1 + 3, y1 + 3, x2 - 3, y2 - 3], outline=(255, 255, 255, 160), width=1)
        cropped, _, _, _, _ = crop_context(img, x1, y1, x2, y2, pad_x=300, pad_y=200)
        col_letter = get_column_letter(target_col)
        st.image(
            cropped,
            use_container_width=True,
            caption=f"Cell {col_letter}{target_row}  ·  Value: {value or '(empty)'}",
        )
    except Exception as e:
        st.error(f"Rendering error: {e}")


# ── Field history dialog ──────────────────────────────────────────────────────

@st.dialog("Field History", width="large")
def show_field_history_dialog(
    field_name: str, sheet: str, claim_id: str,
    current_val: str, original_val: str,
) -> None:
    st.markdown(f"### 📋 History — {field_name}")
    history = _get_field_history(sheet, claim_id, field_name)

    st.markdown(
        f"""
        <div style='background:var(--s0);border:1px solid var(--b0);border-radius:8px;
             padding:12px 16px;margin-bottom:12px;'>
          <div style='display:grid;grid-template-columns:1fr 1fr;gap:16px;'>
            <div>
              <div style='font-size:10px;color:var(--t3);font-family:monospace;text-transform:uppercase;
                   letter-spacing:1px;margin-bottom:6px;'>Original (from file)</div>
              <div style='background:#1a1a2e;border:1px solid #2a2a45;border-radius:5px;
                   padding:8px 12px;font-family:monospace;font-size:13px;color:#f0efff;'>
                {original_val or "(empty)"}
              </div>
            </div>
            <div>
              <div style='font-size:10px;color:var(--t3);font-family:monospace;text-transform:uppercase;
                   letter-spacing:1px;margin-bottom:6px;'>Current Value</div>
              <div style='background:#0f2d1f;border:1px solid rgba(52,211,153,0.35);border-radius:5px;
                   padding:8px 12px;font-family:monospace;font-size:13px;color:#34d399;'>
                {current_val or "(empty)"}
              </div>
            </div>
          </div>
          {"<div style='margin-top:8px;font-size:11px;color:#f5c842;font-family:monospace;'>⚡ Modified from original</div>" if current_val != original_val else "<div style='margin-top:8px;font-size:11px;color:#34d399;font-family:monospace;'>✓ Unchanged from original</div>"}
        </div>
        """,
        unsafe_allow_html=True,
    )

    if history:
        st.markdown("**Edit Timeline**")
        for h in history:
            arrow_col = "var(--yellow)" if h["source"] == "user" else "var(--blue)"
            src_icon  = "✏" if h["source"] == "user" else "⚡"
            src_lbl   = "Manual edit" if h["source"] == "user" else "Auto (LLM/normalize)"
            st.markdown(
                f"<div style='display:flex;align-items:flex-start;gap:12px;padding:10px 0;"
                f"border-bottom:1px solid #1e1e32;'>"
                f"<div style='font-size:10px;color:var(--t3);font-family:monospace;"
                f"white-space:nowrap;margin-top:2px;'>{h['ts']}</div>"
                f"<div style='flex:1;'>"
                f"<div style='display:flex;align-items:center;gap:8px;margin-bottom:6px;'>"
                f"<span style='color:{arrow_col};font-size:12px;'>{src_icon}</span>"
                f"<span style='font-size:11px;color:var(--t3);font-family:monospace;'>{src_lbl}</span></div>"
                f"<div style='display:flex;align-items:center;gap:8px;'>"
                f"<code style='background:#1a1a2e;padding:3px 8px;border-radius:4px;font-size:12px;color:#f0efff;'>"
                f"{h['from'] or '(empty)'}</code>"
                f"<span style='color:{arrow_col};font-size:14px;'>→</span>"
                f"<code style='background:#0f2d1f;padding:3px 8px;border-radius:4px;font-size:12px;color:#34d399;'>"
                f"{h['to'] or '(empty)'}</code></div></div></div>",
                unsafe_allow_html=True,
            )
    else:
        st.markdown(
            "<div style='color:var(--t3);font-size:13px;padding:12px 0;'>"
            "No edits recorded yet for this field.</div>",
            unsafe_allow_html=True,
        )

    if st.button("Close", type="primary", use_container_width=True):
        st.rerun()


# ── Settings dialog ───────────────────────────────────────────────────────────

@st.dialog("Settings", width="large")
def show_settings_dialog(schemas: dict, config_load_status: dict) -> None:
    import os
    from config.settings import CONFIG_DIR

    st.markdown("### Configuration")
    st.markdown("---")
    st.markdown("#### Confidence Settings")

    use_conf = st.checkbox(
        "Enable confidence scoring display",
        value=st.session_state.get("use_conf_threshold", False),
        key="use_conf_toggle",
        help="When enabled, shows confidence scores for each mapped field",
    )
    st.session_state["use_conf_threshold"] = use_conf

    if use_conf:
        conf = st.slider(
            "Confidence threshold", 0, 100,
            value=st.session_state.get("conf_threshold", 80),
            step=5, format="%d%%",
        )
        st.session_state["conf_threshold"] = conf
        bar_color = "#22c55e" if conf >= 70 else "#f59e0b" if conf >= 40 else "#ef4444"
        level_txt = (
            "High confidence — minimal manual review needed" if conf >= 70 else
            "Medium — review flagged fields carefully" if conf >= 40 else
            "Low — most fields will require manual review"
        )
        st.markdown(
            f"<div class='conf-bar-wrap'><div class='conf-bar-fill' "
            f"style='width:{conf}%;background:{bar_color};'></div></div>"
            f"<div style='color:{bar_color};font-size:12px;margin-top:5px;'>{level_txt}</div>",
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            "<div style='color:var(--t3);font-size:13px;font-family:monospace;'>"
            "Confidence scoring is disabled. Enable above to show scores and set threshold.</div>",
            unsafe_allow_html=True,
        )

    st.markdown("---")
    st.markdown("#### Export Schema")
    active_schema = st.session_state.get("active_schema", None)

    for schema_name, schema_def in schemas.items():
        is_active   = active_schema == schema_name
        border_col  = schema_def["color"] if is_active else "#2a2a3e"
        bg_col      = "#1a1a2e" if is_active else "#16161e"
        active_tag  = (
            f"<span style='font-size:10px;color:{schema_def['color']};margin-left:8px;font-weight:bold;'>● ACTIVE</span>"
            if is_active else ""
        )
        custom_count = len(st.session_state.get(f"custom_fields_{schema_name}", []))
        st.markdown(
            f"<div style='background:{bg_col};border:1px solid {border_col};border-radius:8px;"
            f"padding:12px 14px;margin-bottom:4px;'>"
            f"<div style='display:flex;align-items:center;'>"
            f"<span style='font-size:var(--sz-body);font-weight:700;color:var(--t0);font-family:var(--font);'>"
            f"{schema_def['icon']} {schema_name}</span>"
            f"<span style='font-size:var(--sz-sm);color:var(--t3);margin-left:8px;font-family:var(--font);'>"
            f"{schema_def['version']}</span>{active_tag}</div>"
            f"<div style='font-size:var(--sz-sm);color:var(--t2);margin-top:4px;font-family:var(--font);'>"
            f"{schema_def['description']}</div></div>",
            unsafe_allow_html=True,
        )
        bc1, bc2, bc3 = st.columns([1, 1, 1])
        with bc1:
            if st.button(
                "✓ Deactivate" if is_active else "Activate",
                key=f"activate_{schema_name}", use_container_width=True,
            ):
                st.session_state["active_schema"] = None if is_active else schema_name
                st.rerun()
        with bc2:
            if st.button("View Fields", key=f"view_{schema_name}", use_container_width=True):
                st.session_state["schema_popup_target"] = schema_name
                st.session_state["schema_popup_tab"]    = "required"
                st.rerun()
        with bc3:
            if st.button(
                f"Custom Fields ({custom_count})",
                key=f"custom_{schema_name}", use_container_width=True,
            ):
                st.session_state["schema_popup_target"] = schema_name
                st.session_state["schema_popup_tab"]    = "custom"
                st.rerun()
        st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("#### 📁 YAML Config Files")
    st.markdown(
        f"<div style='color:var(--t2);font-size:var(--sz-body);margin-bottom:10px;font-family:var(--font);'>"
        f"Config directory: <code>{CONFIG_DIR}</code></div>",
        unsafe_allow_html=True,
    )
    for schema_name, status in config_load_status.items():
        sc      = schemas.get(schema_name, {})
        col_st  = sc.get("color", "#64748b")
        badge   = (
            "<span style='background:#0f2d1f;border:1px solid #22c55e;border-radius:4px;"
            "padding:1px 7px;font-size:10px;color:#22c55e;'>✓ Loaded</span>"
            if status["loaded"]
            else
            "<span style='background:#2d0f0f;border:1px solid #ef4444;border-radius:4px;"
            "padding:1px 7px;font-size:10px;color:#ef4444;'>✗ Not found — using defaults</span>"
        )
        st.markdown(
            f"<div style='background:var(--s0);border:1px solid var(--b0);border-radius:6px;"
            f"padding:10px 14px;margin-bottom:6px;'>"
            f"<div style='display:flex;align-items:center;gap:10px;'>"
            f"<span style='color:{col_st};font-weight:700;font-size:var(--sz-body);font-family:var(--font);'>"
            f"{sc.get('icon','')} {schema_name}</span>{badge}</div>"
            f"<div style='font-size:var(--sz-xs);color:var(--t3);margin-top:4px;font-family:var(--font);'>"
            f"📄 {status['file']}</div></div>",
            unsafe_allow_html=True,
        )

    if st.button("🔄 Reload YAML Configs", use_container_width=True, key="reload_yaml_cfg"):
        from config.schemas import _load_all_configs, _HARDCODED_SCHEMAS
        import config.schemas as _cs
        _cs.SCHEMAS = _load_all_configs(_HARDCODED_SCHEMAS)
        st.session_state["sheet_cache"] = {}
        st.success("✅ Configs reloaded")
        st.rerun()

    st.markdown("---")
    r1, r2 = st.columns(2)
    with r1:
        if st.button("🗑 Clear Claim Dup History", use_container_width=True, key="clear_claim_dup_all"):
            from modules.claim_dup_store import clear_claim_dup_store
            clear_claim_dup_store()
            st.success("✅ Claim duplicate history cleared")
            st.rerun()
        if st.button("Reset Defaults", use_container_width=True, key="reset_defaults_btn"):
            st.session_state["conf_threshold"]     = 80
            st.session_state["use_conf_threshold"] = False
            st.session_state["active_schema"]      = None
            for s in schemas:
                st.session_state[f"custom_fields_{s}"] = []
            st.rerun()
    with r2:
        if st.button("Close", type="primary", use_container_width=True):
            st.rerun()


# ── Schema field manager dialog ───────────────────────────────────────────────

@st.dialog("Schema Field Manager", width="large")
def show_schema_fields_dialog(schema_name: str, schemas: dict) -> None:
    schema     = schemas[schema_name]
    custom_key = f"custom_fields_{schema_name}"
    if custom_key not in st.session_state:
        st.session_state[custom_key] = []

    st.markdown(f"### {schema['icon']} {schema_name} — {schema['version']}")
    st.markdown(
        f"<div style='color:var(--t2);font-size:var(--sz-body);margin-bottom:14px;font-family:var(--font);'>"
        f"{schema['description']}</div>",
        unsafe_allow_html=True,
    )
    tab_req, tab_accepted, tab_custom = st.tabs(["Mandatory Fields", "All Accepted Fields", "My Custom Fields"])

    with tab_req:
        pills = "".join(
            f"<span class='field-pill field-pill-required'>✓ {f}</span>"
            for f in schema["required_fields"]
        )
        st.markdown(f"<div style='margin:12px 0;'>{pills}</div>", unsafe_allow_html=True)

    with tab_accepted:
        optional  = [f for f in schema["accepted_fields"] if f not in schema["required_fields"]]
        req_pills = "".join(f"<span class='field-pill field-pill-required'>✓ {f}</span>" for f in schema["required_fields"])
        opt_pills = "".join(f"<span class='field-pill'>{f}</span>" for f in optional)
        st.markdown(
            f"<div style='margin:12px 0;'>"
            f"<b style='color:var(--t2);font-size:var(--sz-xs);font-family:var(--font);letter-spacing:1.2px;"
            f"text-transform:uppercase;'>MANDATORY</b>"
            f"<br><div style='margin-top:6px;'>{req_pills}</div></div>"
            f"<div style='margin:12px 0;'>"
            f"<b style='color:var(--t2);font-size:var(--sz-xs);font-family:var(--font);letter-spacing:1.2px;"
            f"text-transform:uppercase;'>OPTIONAL</b>"
            f"<br><div style='margin-top:6px;'>{opt_pills}</div></div>",
            unsafe_allow_html=True,
        )

    with tab_custom:
        custom_fields = st.session_state[custom_key]
        already_added = set(custom_fields) | set(schema["required_fields"])
        available     = [f for f in schema["accepted_fields"] if f not in already_added]

        if available:
            sel_col, add_col = st.columns([4, 1])
            with sel_col:
                chosen = st.selectbox(
                    "Pick field",
                    ["— select a field —"] + available,
                    key=f"new_field_sel_{schema_name}",
                    label_visibility="collapsed",
                )
            with add_col:
                if st.button("Add", key=f"add_field_btn_{schema_name}", use_container_width=True, type="primary"):
                    if chosen and chosen != "— select a field —":
                        st.session_state[custom_key].append(chosen)
                        st.rerun()

        if not custom_fields:
            st.markdown(
                "<div style='color:var(--t2);font-size:var(--sz-body);padding:10px 0;font-family:var(--font);'>"
                "No optional fields added yet.</div>",
                unsafe_allow_html=True,
            )
        else:
            for idx, cf in enumerate(list(custom_fields)):
                cf1, cf2 = st.columns([5, 1])
                with cf1:
                    cls = "field-pill-required" if cf in schema["required_fields"] else "field-pill-custom"
                    icon = "✓" if cf in schema["required_fields"] else "+"
                    st.markdown(f"<span class='field-pill {cls}'>{icon} {cf}</span>", unsafe_allow_html=True)
                with cf2:
                    if st.button("Remove", key=f"del_cf_{schema_name}_{idx}", use_container_width=True):
                        st.session_state[custom_key].pop(idx)
                        st.rerun()
            st.markdown("---")
            if st.button(f"Clear All", key=f"clear_all_{schema_name}"):
                st.session_state[custom_key] = []
                st.rerun()

        total = len(schema["required_fields"]) + len(custom_fields)
        st.markdown(
            f"<div style='background:var(--s0);border:1px solid var(--b0);border-radius:8px;"
            f"padding:10px 16px;'>"
            f"<span style='color:var(--t2);font-size:var(--sz-body);font-family:var(--font);'>"
            f"Mandatory: <b style='color:var(--blue);'>{len(schema['required_fields'])}</b>"
            f" &nbsp;|&nbsp; Custom: <b style='color:var(--green);'>{len(custom_fields)}</b>"
            f" &nbsp;|&nbsp; Total: <b style='color:var(--t0);'>{total}</b></span></div>",
            unsafe_allow_html=True,
        )
