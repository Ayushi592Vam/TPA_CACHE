"""
modules/parsing.py
Excel / CSV ingestion: classify sheet type, parse rows into list-of-dicts,
skip aggregate/totals rows.
"""

import csv
import os
import re

import openpyxl

from modules.cell_format import format_cell_value_with_fmt


# ── Sheet classifier ──────────────────────────────────────────────────────────

def classify_sheet(rows) -> str:
    text = " ".join(str(cell).lower() for row in rows[:20] for cell in row if cell)
    if "line of business" in text:
        return "SUMMARY"
    has_claim = any(x in text for x in [
        "claim number", "claim no", "claim #", "claim id", "claim_id",
        "claim ref", "claimant", "file number", "file no",
    ])
    has_loss = any(x in text for x in [
        "loss date", "date of loss", "loss dt", "accident date",
        "occurrence date", "incident date", "date of injury", "date of incident",
        "injury date",
    ])
    has_fin = any(x in text for x in [
        "incurred", "paid", "reserve", "outstanding",
        "total paid", "total incurred", "indemnity", "expense",
    ])
    if has_claim and (has_loss or has_fin):
        return "LOSS_RUN"
    if "policy" in text and ("claim" in text or "incurred" in text):
        return "COMMERCIAL_LOSS_RUN"
    if has_claim:
        return "LOSS_RUN"
    return "UNKNOWN"


# ── Aggregate-row detection ───────────────────────────────────────────────────

_AGGREGATE_PATTERNS = re.compile(
    r"^(total|totals|grand\s*total|subtotal|aggregate|summary|sum|report\s*(date|total|summary)|"
    r"all\s+adjusters|ytd\s+total|period\s+total|fiscal\s+total|portfolio\s+total|"
    r"TOTALS_AGGREGATE|SUMMARY_FLIBBER|AGGREGATE_ZORP|SUMMARY_ZORP)",
    re.IGNORECASE,
)
_AGGREGATE_EXTRA = re.compile(
    r"(aggregate|zorp|flibber|summary|zoop|gorp|totals?_|_total|report_date|all_adjuster)",
    re.IGNORECASE,
)


def _is_aggregate_row(row_values: list) -> bool:
    non_empty = [str(v).strip() for v in row_values if v is not None and str(v).strip()]
    if not non_empty:
        return False
    first_val = non_empty[0]
    if _AGGREGATE_PATTERNS.match(first_val):
        return True
    if _AGGREGATE_EXTRA.search(first_val):
        return True
    first_tokens    = re.split(r"[_\s]+", first_val.lower())
    aggregate_tokens = {"total", "totals", "aggregate", "summary", "subtotal", "grand", "portfolio", "report"}
    if len(first_tokens) >= 2 and any(t in aggregate_tokens for t in first_tokens):
        return True
    for v in non_empty[:6]:
        if re.match(
            r"(total\s+claims|report\s+date|all\s+adjusters|open:\s*\d|pend:\s*\d|open:\d)",
            str(v), re.IGNORECASE,
        ):
            return True
    nums = [float(v) for v in row_values if isinstance(v, (int, float))]
    if nums and len(nums) >= 3 and all(n > 50_000 for n in nums):
        if not re.match(r"^[A-Z]{2,5}[-_][A-Z]{0,3}\d{3,}", first_val, re.IGNORECASE):
            return True
    return False


# ── Main entry point ──────────────────────────────────────────────────────────

def extract_from_excel(file_path: str, sheet_name: str) -> tuple[list, str]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        if not rows:
            return [], "UNKNOWN"
        return parse_rows(classify_sheet(rows), rows)
    else:
        wb        = openpyxl.load_workbook(file_path, data_only=True)
        ws        = wb[sheet_name]
        raw_rows  = [[cell.value for cell in row] for row in ws.iter_rows()]
        cell_rows = [list(row) for row in ws.iter_rows()]
        wb.close()
        if not raw_rows:
            return [], "UNKNOWN"
        return parse_rows_with_cells(classify_sheet(raw_rows), raw_rows, cell_rows)


# ── Row parsers ───────────────────────────────────────────────────────────────

def _find_header_row(rows: list) -> int | None:
    for i, row in enumerate(rows[:20]):
        rt = " ".join([str(c).lower() for c in row if c])
        if (
            "claim" in rt or "employee name" in rt or "driver name" in rt
            or "claimant" in rt or "file" in rt
        ) and (
            "date" in rt or "incurred" in rt or "paid" in rt
            or "injury" in rt or "incident" in rt or "amount" in rt or "reserve" in rt
        ):
            return i
    # Fallback
    for i, row in enumerate(rows[:5]):
        if sum(1 for c in row if c) >= 3:
            return i
    return None


def parse_rows_with_cells(sheet_type: str, rows: list, cell_rows: list) -> tuple[list, str]:
    if sheet_type == "SUMMARY":
        hri = None
        for i, row in enumerate(rows[:20]):
            rt = " ".join([str(c).lower() for c in row if c])
            if "sheet" in rt and "line of business" in rt:
                hri = i
                break
        if hri is None:
            return [], sheet_type
        headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        extracted = []
        for r_idx_rel, (raw_row, cell_row) in enumerate(zip(rows[hri + 1:], cell_rows[hri + 1:])):
            r_idx = hri + 2 + r_idx_rel
            if not any(raw_row):
                continue
            row_data: dict = {}
            for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
                if c_idx_0 >= len(headers):
                    continue
                clean_val = format_cell_value_with_fmt(cell)
                real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
                row_data[headers[c_idx_0]] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": real_col,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    hri = _find_header_row(rows)
    if hri is None:
        return [], sheet_type
    headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
    extracted = []
    for r_idx_rel, (raw_row, cell_row) in enumerate(zip(rows[hri + 1:], cell_rows[hri + 1:])):
        r_idx = hri + 2 + r_idx_rel
        if not any(raw_row):
            continue
        if any(str(c).lower().strip() in ["totals", "total", "grand total", "subtotal"] for c in raw_row if c):
            break
        if _is_aggregate_row(raw_row):
            continue
        row_data: dict = {}
        for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
            if c_idx_0 >= len(headers):
                continue
            clean_val = format_cell_value_with_fmt(cell)
            real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
            row_data[headers[c_idx_0]] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": real_col,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


def parse_rows(sheet_type: str, rows: list) -> tuple[list, str]:
    if sheet_type == "SUMMARY":
        hri = None
        for i, row in enumerate(rows[:20]):
            rt = " ".join([str(c).lower() for c in row if c])
            if "sheet" in rt and "line of business" in rt:
                hri = i
                break
        if hri is None:
            return [], sheet_type
        headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        extracted = []
        for r_idx, row in enumerate(rows[hri + 1:], start=hri + 2):
            if not any(row):
                continue
            if _is_aggregate_row(list(row)):
                continue
            row_data: dict = {}
            for c_idx, value in enumerate(row, start=1):
                if c_idx - 1 >= len(headers):
                    continue
                clean_val = str(value).strip() if value is not None else ""
                row_data[headers[c_idx - 1]] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": c_idx,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    hri = _find_header_row(rows)
    if hri is None:
        return [], sheet_type
    headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
    extracted = []
    for r_idx, row in enumerate(rows[hri + 1:], start=hri + 2):
        if not any(row):
            continue
        if any(str(cell).lower().strip() in ["totals", "total", "grand total"] for cell in row if cell):
            break
        if _is_aggregate_row(list(row)):
            continue
        row_data: dict = {}
        for c_idx, value in enumerate(row, start=1):
            if c_idx - 1 >= len(headers):
                continue
            clean_val = str(value).strip() if value is not None else ""
            row_data[headers[c_idx - 1]] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": c_idx,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type
